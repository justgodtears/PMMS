import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

def process_ptv_data(df):
    """
    Przetwarza dane zgodnie z logiką VBA
    """
    # Usuń pierwszy wiersz (stare nagłówki) - VBA zaczyna od wiersza 1 który jest nagłówkiem
    # W pandas pierwszy wiersz to indeks 0
    df_without_headers = df.iloc[1:].reset_index(drop=True)
    
    # Kolumna "Hoja Ruta" (indeks 46, bo Python od 0)
    hoja_ruta_col = 46
    
    # Konwertuj kolumnę na numeric (obsługa stringów i pustych wartości)
    hoja_ruta_series = pd.to_numeric(df_without_headers.iloc[:, hoja_ruta_col], errors='coerce')
    
    # Filtruj wiersze gdzie Hoja Ruta <= 0 lub pusty (NaN po konwersji)
    mask = (hoja_ruta_series.isna()) | (hoja_ruta_series <= 0)
    filtered_df = df_without_headers[mask].copy()
    
    # --- Definicja kolumn ---
    # Stare kolumny (indeksy Python).
    # "penalty cost" (dawny index 42) jest usuwany.
    # Kolumny, które miały target > 42, muszą mieć zmniejszony indeks o 1.
    
    # Lista par (Source Index, Target Index)
    # Target Index to 0-based indeks w result_df
    
    # Stare mapowania (zaktualizowane o przesunięcie):
    # Old Target: [47, 26, 3, 4, 5, 7, 6, 48, 49, 20, 19, 21, 22, 23, 24, 25]
    # Source:     [3,  8, 15, 10, 16, 17, 18, 21, 29, 35, 36, 37, 38, 39, 40, 54]
    
    mappings = [
        (3, 47),   # id -> avis (Old 47 -> New 47? Wait. Old 47 was "avis"? No.)
                   # Let's verify old headers/indices again.
                   # Old Headers:
                   # 42: penalty cost (REMOVED)
                   # 43: group id (New 42)
                   # 44: same vehicle group id (New 43)
                   # 45: sequence number (New 44)
                   # 46: sequence group id (New 45)
                   # 47: avis (New 46)
                   # 48: avis pickup date (New 47)
                   # 49: final destination (New 48)
        
        # Checking logic of Old Script:
        # cols_to_keep = [3, 8, ...]
        # target_columns = [47, 26, ...]
        # So Source 3 went to Target 47 ("avis"). 
        # Source 21 went to Target 48 ("avis pickup date").
        # Source 29 went to Target 49 ("final destination").
        
        # New Targets (since 42 removed):
        # Old 47 -> New 46
        # Old 48 -> New 47
        # Old 49 -> New 48
         
        (3, 46),   # id -> avis
        (8, 26),   # linked order id -> stacking factor
        (15, 3),   # location id -> location id
        (10, 4),   # location name -> location name
        (16, 5),   # location street -> location street
        (17, 7),   # location zip code -> location city (Old 7)?
                   # Wait, Old Tgt 7 was "location city". Old Src 17. 
                   # Old 6 was "location zip code". Old Src 18.
                   # Let's stick to the code:
                   # Src 16 -> Tgt 5 ("location street")
                   # Src 17 -> Tgt 7 ("location city")
                   # Src 18 -> Tgt 6 ("location zip code")
        (18, 6),
        (21, 47),  # avis pickup date
        (29, 48),  # final destination
        (35, 20),  # weight
        (36, 19),  # volume
        (37, 21),  # Height
        (38, 22),  # Width
        (39, 23),  # Lenght
        (40, 24),  # pc
        (54, 25)   # support data
    ]
    
    # Nowe kolumny mapowane z pliku
    # 49: destination id (Src 26)
    # 50: destination name (Src 28)
    # 51: destination street (Src 34)
    # 52: destination zip code (Src 27)
    # 53: destination city (Src 30) - CLEAN "INT"
    
    new_mappings = [
        (25, 49),
        (27, 50),
        (33, 51),
        (26, 52),
        (29, 53)
    ]
    
    # Scal wszystkie mapowania
    all_mappings = mappings + new_mappings
    
    # Nowe nagłówki (63 kolumny)
    headers = [
        "id", "linked order id", "type", "location id",
        "location name", "location street", "location zip code",
        "location city", "location country", "location group stop time",
        "location stop time", "latitude", "longitude", "loading meters",
        "Capacity 1", "Capacity 2", "Capacity 3", "Capacity 4", "Capacity 5",
        "volume", "weight", "Height", "Width", "Lenght", "pc",
        "support data for stacking factor", "stacking factor",
        "Corrected stacking factor", "Corrected stacking factor 1",
        "Corrected stacking factor 2", "Corrected stacking factor 3",
        "Corrected stacking factor 4", "Corrected stacking factor 5",
        "Corrected stacking factor 6", "service time", "absolute timewindows",
        "time window type", "color", "as is sequence", "tags",
        "forbidden tags", "labels", 
        # "penalty cost" removed
        "group id", "same vehicle group id", "sequence number", 
        "sequence group id", "avis", "avis pickup date", "final destination",
        # New Columns
        "destination id", "destination name", "destination street",
        "destination zip code", "destination city", "destination country",
        "destination latitude", "destination longitude", 
        "destination absolute timewindows", "weight avis", 
        "loading meter avis", "max avis", "penalty cost", "location counted as stop"
    ]
    
    # Utwórz pustą DataFrame z 63 kolumnami
    result_df = pd.DataFrame(index=range(len(filtered_df)), columns=range(63))
    
    # Przepisz dane
    for src_col, tgt_col in all_mappings:
        if src_col < len(filtered_df.columns):
            vals = filtered_df.iloc[:, src_col].values
            
            # Czyszczenie "INT" dla Destination City (Tgt 53)
            if tgt_col == 53:
                vals = [str(x).replace("INT - ", "").strip() if pd.notna(x) else x for x in vals]
            
            result_df[tgt_col] = vals
            
    # Ustaw nagłówki
    result_df.columns = headers
    
    # Wypełnij stałe wartości
    result_df['type'] = 'PICKUP'
    result_df['location country'] = 'DE'
    result_df['destination country'] = 'DE' # Nowa stała
    result_df['location stop time'] = '1:00'
    result_df['service time'] = '0:01'
    result_df['time window type'] = 'Service Arrival'
    result_df['location counted as stop'] = "True" # Nowa stała
    result_df['destination absolute timewindows'] = "06:00:00-16:00:00" # Nowa stała
    
    return result_df

def add_formulas_to_excel(df, output_path):
    """
    Dodaje formuły Excel do pliku wyjściowego
    """
    # Zapisz DataFrame do Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Formatowanie: Kolorowanie volume > 99
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF')
        
        oversized_rows = []
        
        # Iteruj po wierszach danych (od 2)
        for i in range(2, len(df) + 2):
            row = i
            
            # --- Stare formuły (sprawdzone indeksy) ---
            # absolute timewindows (Col 35 -> AJ). Use G (Col 6).
            worksheet[f'AJ{row}'].value = f"=IFERROR(VLOOKUP(G{row},'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[opentime.xlsx]sheet'!$B$2:$F$1000,5,FALSE),\"06:00:00-14:00:00\")"
            
            # Latitude (Col 11 -> L). Use E.
            worksheet[f'L{row}'] = f'=IFERROR(VLOOKUP(E{row},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[Points of loading PTV.xlsx]ptvform\'!$A:$D,2,FALSE),"Nie znaleziono")'
            
            # Longitude (Col 12 -> M). Use E.
            worksheet[f'M{row}'] = f'=IFERROR(VLOOKUP(E{row},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[Points of loading PTV.xlsx]ptvform\'!$A:$D,3,FALSE),"Nie znaleziono")'
            
            # Corrected stacking factors (AA to AH)
            worksheet[f'AB{row}'] = f'=AA{row}+1'
            worksheet[f'AC{row}'] = f'=IF(INT(3/V{row}) > AB{row}, AB{row}, INT(3/V{row}))'
            worksheet[f'AD{row}'] = f'=IF(INT(2.7/V{row}) > AB{row}, AB{row}, INT(2.7/V{row}))'
            worksheet[f'AE{row}'] = f'=IF(INT(2.7/V{row}) > AB{row}, AB{row}, INT(2.7/V{row}))'
            worksheet[f'AF{row}'] = f'=IF(INT(2.34/V{row}) > AB{row}, AB{row}, INT(2.34/V{row}))'
            worksheet[f'AG{row}'] = f'=IF(INT(2.3/V{row}) > AB{row}, AB{row}, INT(2.3/V{row}))'
            worksheet[f'AH{row}'] = f'=IF(INT(1.86/V{row}) > AB{row}, AB{row}, INT(1.86/V{row}))'
            
            # Capacities (N to S)
            worksheet[f'N{row}'] = f'=IF(OR(AC{row}=0, X{row}>13.6, W{row}>2.48), 999, ROUNDUP(Y{row}/AC{row},0) * ((W{row}*X{row})/2.48))'
            worksheet[f'O{row}'] = f'=IF(OR(AD{row}=0, X{row}>13.6, W{row}>2.48), 999, ROUNDUP(Y{row}/AD{row},0) * ((W{row}*X{row})/2.48))'
            worksheet[f'P{row}'] = f'=IF(OR(AE{row}=0, X{row}>7.2, W{row}>2.48), 999, ROUNDUP(Y{row}/AE{row},0) * ((W{row}*X{row})/2.48))'
            worksheet[f'Q{row}'] = f'=IF(OR(AF{row}=0, X{row}>7.2, W{row}>2.48), 999, ROUNDUP(Y{row}/AF{row},0) * ((W{row}*X{row})/2.48))'
            worksheet[f'R{row}'] = f'=IF(OR(AG{row}=0, X{row}>4.2, W{row}>2.2), 999, ROUNDUP(Y{row}/AG{row},0) * ((W{row}*X{row})/2.2))'
            worksheet[f'S{row}'] = f'=IF(OR(AH{row}=0, X{row}>4, W{row}>1.56), 999, ROUNDUP(Y{row}/AH{row},0) * ((W{row}*X{row})/1.56))'
            
            # tags (Col 39 -> AN)
            worksheet[f'AN{row}'] = f'=IF(E{row}="LEM EUROPE GMBH","ML",IF(E{row}="IEC GMBH","ML",IF(E{row}="ANTARES LIFE CYCLE SOLUTION GMBH","SM",IF(E{row}="HEIMSCH DESIGN GMBH","NOBGM",IF(E{row}="HENKEL WERK HEIDELBERG","ADR",IF(E{row}="PROMENS (ROTOVIA) HOCKENHEIM GMBH","ROT",""))))))'
            
            # labels (Col 41 -> AP)
            worksheet[f'AP{row}'] = f'=IF(LEFT(SUBSTITUTE(G{row},"DE-",""),1)="5","AREA 1",IF(LEFT(SUBSTITUTE(G{row},"DE-",""),1)="6","AREA 2",IF(LEFT(SUBSTITUTE(G{row},"DE-",""),1)="7","AREA 3","")))'
            
            # color (Col 37 -> AL)
            worksheet[f'AL{row}'] = f'=IF(AP{row}="AREA 1","blue",IF(AP{row}="AREA 2","green",IF(AP{row}="AREA 3","yellow","")))'
            
            # group id (Col 42 -> AQ). Was AS previously. Reference New "same vehicle id"?
            # Wait. Logic check.
            # Old: AS = AV. (Same Veh = Avis).
            # New: AR (Same Veh, Col 43) = AU (Avis, Col 46).
            worksheet[f'AR{row}'] = f'=AU{row}'
            
            # --- Nowe Formuły ---

            # destination latitude 
            worksheet[f'BD{row}'] = f'=IFERROR(VLOOKUP(AY{row},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[direct_flows.xlsx]adress\'!$A:$B,2,FALSE),"Nie znaleziono")'
            
            # destination longitude
            worksheet[f'BE{row}'] = f'=IFERROR(VLOOKUP(AY{row},\'https://juliob50600527-my.sharepoint.com/personal/oliwier_opyrchal_gruposese_com/Documents/[direct_flows.xlsx]adress\'!$A:$C,3,FALSE),"Nie znaleziono")'

            # weight avis (Col 58 -> BG)
            worksheet[f'BG{row}'] = f'=SUMIFS(U:U, AU:AU, AU{row}, C:C, "PICKUP")/24000'
            
            # loading meter avis (Col 59 -> BH)
            worksheet[f'BH{row}'] = f'=SUMIFS(N:N, AU:AU, AU{row}, C:C, "PICKUP")/13.6'
            
            # max avis (Col 60 -> BI)
            worksheet[f'BI{row}'] = f'=MAX(BG{row}, BH{row})'
            
            # penalty cost (Col 61 -> BJ)
            # Używamy angielskich nazw funkcji dla openpyxl
            worksheet[f'BJ{row}'] = (
                f'=IF(AND(BI{row}>=0.75, BI{row}<=1), 10000, '
                f'IF(AND(BI{row}>=0.5, BI{row}<0.75), 7500, '
                f'IF(AND(BI{row}>=0.3, BI{row}<0.5), 5000, '
                f'IF(AND(BI{row}>=0.2, BI{row}<0.3), 2500, '
                f'IF(COUNTIF(AU:AU, AU{row})>2, 1, 0)))))'
            )
            
            # Sprawdzenie volume > 99
            volume_cell = worksheet[f'T{row}']
            if volume_cell.value and isinstance(volume_cell.value, (int, float)):
                if volume_cell.value > 99:
                    oversized_rows.append(row)
                    for col in range(1, 64): # Do 63 kolumny
                        cell = worksheet.cell(row=row, column=col)
                        cell.fill = red_fill
                        cell.font = white_font
        
        # Obramowanie
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=63):
            for cell in row:
                cell.border = thin_border
        
        # AutoFit
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    # Ustawienia obliczeń
    workbook.calculation.calcMode = 'auto'
    workbook.calculation.fullCalcOnLoad = True
    return oversized_rows

# Streamlit App
st.set_page_config(page_title="PTV Data Converter", page_icon="📊", layout="wide")

st.title("🚚 PTV Data Converter with Directs")
st.markdown("---")

st.markdown("""
### Instrukcja:
1. Wgraj plik Excel (wsadowy plik przed użyciem VBA)
2. Kliknij **Convert**
3. Pobierz przekonwertowany plik

**Funkcje skryptu:**
- Filtruje dane gdzie `Hoja Ruta <= 0` lub puste
- Reorganizuje kolumny do formatu PTV (50 kolumn)
- Dodaje formuły VLOOKUP, obliczenia i stałe wartości
- Oznacza czerwonym wiersze z `volume > 99`

            
**Ważne**
Po pobraniu pliku i dodaniu etykiety etc. w kolumnie absolute timewindows trzeba enter w pasku formuły.
Pasek formuły to ten biały pasek z tekstem który znajduje się pod paskiem z narzędziami i tam klikamy enter.
Bez tego formuła się nie wczyta, wynika to z błędu excela.
Po wciśnięciu enter, wystarczy przeciągnąć formułę po wszystkich komórkach. """)


st.markdown("---")

# Upload file
uploaded_file = st.file_uploader("📁 Wybierz plik Excel", type=['xlsx', 'xls'])

if uploaded_file is not None:
    try:
        # Wczytaj plik
        df = pd.read_excel(uploaded_file, header=None)
        
        st.success(f"✅ Wczytano plik: **{uploaded_file.name}**")
        st.info(f"📊 Wierszy: {len(df)} | Kolumn: {len(df.columns)}")
        
        # Pokaż preview
        with st.expander("👁️ Podgląd danych wejściowych (pierwsze 5 wierszy)"):
            st.dataframe(df.head())
        
        # Convert button
        if st.button("🔄 Convert", type="primary", use_container_width=True):
            with st.spinner("Przetwarzanie danych..."):
                # Przetwórz dane
                result_df = process_ptv_data(df)
                
                # Zapisz do tymczasowego pliku z formułami
                output_buffer = BytesIO()
                oversized_rows = add_formulas_to_excel(result_df, output_buffer)
                output_buffer.seek(0)
                
                st.success("✅ Konwersja zakończona!")
                
                # Pokaż statystyki
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Wierszy po filtracji", len(result_df))
                with col2:
                    st.metric("Kolumn wyjściowych", 50)
                with col3:
                    st.metric("Ponadnormatywny ładunek", len(oversized_rows))
                
                # Ostrzeżenie o ponadnormatywnym ładunku
                if oversized_rows:
                    st.warning(f"⚠️ **UWAGA!** Znaleziono ponadnormatywny ładunek (volume > 99) w wierszach: {', '.join(map(str, oversized_rows))}")
                
                # Pokaż preview wyniku
                with st.expander("👁️ Podgląd danych wyjściowych (pierwsze 10 wierszy)"):
                    st.dataframe(result_df.head(10))
                
                # Download button
                st.download_button(
                    label="⬇️ Pobierz przekonwertowany plik",
                    data=output_buffer,
                    file_name=f"converted_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
    except Exception as e:
        st.error(f"❌ Błąd podczas przetwarzania: {str(e)}")
        st.exception(e)

else:
    st.info("👆 Wgraj plik Excel aby rozpocząć konwersję")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; font-size: 12px;'>
    PTV Data Converter | Powered by Python & Streamlit
</div>
""", unsafe_allow_html=True)
